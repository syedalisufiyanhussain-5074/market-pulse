"""Independent Validation Excel generator.

Sheet 1: Forecast Comparison — 2×2 grid (≤12 predictions) or horizontal (>12)
Sheet 2: Metrics Comparison — MAE/SMAPE/MFE comparison + Agreement Score
"""

import io
from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XlImage
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from app.services.excel_export import LOGO_PATH, FREQ_LABELS, _format_date
from app.services.independent_validation import MODEL_ORDER
from app.utils.logger import get_logger

logger = get_logger("independent_validation_export")

# Styling constants
_border_side = Side(style="thin", color="9CA3AF")
_border = Border(top=_border_side, bottom=_border_side, left=_border_side, right=_border_side)
_header_font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
_center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
_data_font = Font(name="Calibri", size=10, color="1F2937")
_data_align = Alignment(horizontal="center", vertical="center")
_number_fmt = "#,##0.00"

# Header color groups
_fill_base = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")       # Original dark
_fill_mp = PatternFill(start_color="248479", end_color="248479", fill_type="solid")          # Deep Teal (darker)
_fill_r = PatternFill(start_color="C3791F", end_color="C3791F", fill_type="solid")           # Burnt Amber (darker)
_fill_analysis = PatternFill(start_color="495057", end_color="495057", fill_type="solid")    # Cool Grey

# Status colors
_green_fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
_yellow_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
_red_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
_green_font = Font(name="Calibri", bold=True, size=10, color="065F46")
_yellow_font = Font(name="Calibri", bold=True, size=10, color="92400E")
_red_font = Font(name="Calibri", bold=True, size=10, color="991B1B")

# Variance highlight (>2%) — matches Excel's "Highlight Duplicate Values" style
_var_highlight_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
_var_highlight_font = Font(name="Calibri", size=10, color="9C0006")

# Display names
_DISPLAY = {
    "AutoETS": "AutoETS",
    "AutoARIMA": "AutoARIMA",
    "Moving Average (Excel)": "Moving Average",
    "ETS (Excel)": "ETS Excel",
}

_DISPLAY_FULL = {
    "AutoETS": "AutoETS",
    "AutoARIMA": "AutoARIMA",
    "Moving Average (Excel)": "Moving Average (Excel)",
    "ETS (Excel)": "ETS (Excel)",
}

_DISPLAY_STATUS = {
    "AutoETS": "AutoETS's",
    "AutoARIMA": "AutoARIMA's",
    "Moving Average (Excel)": "Moving Average (Excel)",
    "ETS (Excel)": "ETS (Excel)",
}

SPACER_WIDTH = 4.0
DATA_COL_WIDTH = 14.0


def _status_style(status: str):
    if status == "Strong":
        return _green_fill, _green_font
    elif status == "Moderate":
        return _yellow_fill, _yellow_font
    return _red_fill, _red_font


def _add_logo(ws):
    if LOGO_PATH.exists():
        logo = XlImage(str(LOGO_PATH))
        logo.width = 150
        logo.height = int(150 * 791 / 2160)
        ws.add_image(logo, "A1")
    ws.row_dimensions[1].height = 20.55
    ws.row_dimensions[2].height = 20.55


def _write_header_cell(ws, row, col, value, fill):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = _header_font
    cell.fill = fill
    cell.alignment = _center_align
    cell.border = _border


def _write_data_cell(ws, row, col, value, fmt=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = _data_font
    cell.alignment = _data_align
    cell.border = _border
    if fmt:
        cell.number_format = fmt


def _write_status_cell(ws, row, col, status):
    cell = ws.cell(row=row, column=col, value=status)
    fill, font = _status_style(status)
    cell.font = font
    cell.fill = fill
    cell.alignment = _data_align
    cell.border = _border


def _write_model_group_headers(ws, row, start_col, model_name, selected_model):
    d = _DISPLAY[model_name]
    mp_label = f"MP-{d} (Primary Model)" if model_name == selected_model else f"MP-{d}"
    headers = [
        ("Date", _fill_base),
        (mp_label, _fill_mp),
        (f"R-{d} (Validation)", _fill_r),
        (f"{_DISPLAY_FULL[model_name]} Variance", _fill_analysis),
        (f"{_DISPLAY_FULL[model_name]} Variance %", _fill_analysis),
        (f"{_DISPLAY_STATUS[model_name]} Status", _fill_analysis),
    ]
    for i, (label, fill) in enumerate(headers):
        _write_header_cell(ws, row, start_col + i, label, fill)
    # Left-align Date header
    ws.cell(row=row, column=start_col).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)


def _write_model_group_data(ws, row, start_col, model_name, idx, forecast_data,
                            py_forecasts, ind_forecasts, variance_data, frequency):
    date_str = _format_date(forecast_data[idx]["date"], frequency)
    _write_data_cell(ws, row, start_col, date_str)
    # Left-align Date data
    ws.cell(row=row, column=start_col).alignment = Alignment(horizontal="left", vertical="center")

    py_vals = py_forecasts.get(model_name, [])
    _write_data_cell(ws, row, start_col + 1,
                     py_vals[idx] if idx < len(py_vals) else None, _number_fmt)

    ind_vals = ind_forecasts.get(model_name, [])
    _write_data_cell(ws, row, start_col + 2,
                     ind_vals[idx] if idx < len(ind_vals) else None, _number_fmt)

    var_vals = variance_data.get(model_name, {}).get("var", [])
    var_val = var_vals[idx] if idx < len(var_vals) else None
    _write_data_cell(ws, row, start_col + 3, var_val, _number_fmt)

    pct_vals = variance_data.get(model_name, {}).get("var_pct", [])
    pct_val = pct_vals[idx] if idx < len(pct_vals) else None
    _write_data_cell(ws, row, start_col + 4, pct_val, '0.00"%"')

    # Per-prediction status + color coding (Variance, Variance %, Status all in sync)
    if pct_val is not None:
        abs_pct = abs(pct_val)
        if abs_pct <= 5:
            fill, font, status = _green_fill, _green_font, "Strong"
        elif abs_pct <= 10:
            fill, font, status = _yellow_fill, _yellow_font, "Moderate"
        else:
            fill, font, status = _var_highlight_fill, _var_highlight_font, "Weak"
        for col_offset in (3, 4):
            cell = ws.cell(row=row, column=start_col + col_offset)
            cell.fill = fill
            cell.font = font
    else:
        status = ""

    _write_status_cell(ws, row, start_col + 5, status)


def _set_spacer(ws, col):
    ws.column_dimensions[get_column_letter(col)].width = SPACER_WIDTH


def _autofit_group(ws, start_col, count=6):
    for i in range(count):
        ws.column_dimensions[get_column_letter(start_col + i)].width = DATA_COL_WIDTH


def generate_independent_validation_excel(
    historical_data: list[dict],
    forecast_data: list[dict],
    ind_forecasts: dict[str, list[float]],
    py_forecasts: dict[str, list[float]],
    variance_data: dict[str, dict],
    ind_metrics: dict[str, dict],
    py_metrics: dict[str, dict],
    agreement_score: float | None,
    frequency: str,
    selected_model: str = "AutoETS",
    validation_warnings: list[str] | None = None,
) -> bytes:
    freq_label = FREQ_LABELS.get(frequency, frequency)
    fc_count = len(forecast_data)
    hist_count = len(historical_data)
    wb = Workbook()

    # ── Sheet 1: Forecast Comparison ─────────────────────────────────
    ws1 = wb.active
    ws1.title = "Forecast Comparison"
    ws1.sheet_view.showGridLines = False
    ws1.sheet_view.zoomScale = 100

    _add_logo(ws1)

    title_font = Font(name="Calibri", bold=True, size=14, color="1F2937")
    ws1["A3"] = f"Market Pulse \u2014 Independent Validation ({freq_label})"
    ws1["A3"].font = title_font
    ws1["A3"].alignment = Alignment(horizontal="left", vertical="center")
    ws1.row_dimensions[3].height = 20.55

    ws1.column_dimensions["A"].width = DATA_COL_WIDTH
    ws1.column_dimensions["B"].width = DATA_COL_WIDTH

    use_grid = fc_count <= 12
    models = list(MODEL_ORDER)

    # ── Actuals table (A-B): only historical data, no forecast dates ──
    ws1.row_dimensions[4].height = 27.6
    _write_header_cell(ws1, 4, 1, "Date", _fill_base)
    ws1.cell(row=4, column=1).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    _write_header_cell(ws1, 4, 2, "Actual", _fill_base)

    for i, entry in enumerate(historical_data):
        row = 5 + i
        date_str = _format_date(entry["date"], frequency)
        _write_data_cell(ws1, row, 1, date_str)
        ws1.cell(row=row, column=1).alignment = Alignment(horizontal="left", vertical="center")
        _write_data_cell(ws1, row, 2, entry["value"], _number_fmt)

    if use_grid:
        # ── 2×2 Grid Layout ──
        _set_spacer(ws1, 3)   # C
        _set_spacer(ws1, 10)  # J
        group1_col = 4   # D
        group2_col = 11  # K
        _autofit_group(ws1, group1_col)
        _autofit_group(ws1, group2_col)

        # Top two models: headers at row 4, data starts row 5
        _write_model_group_headers(ws1, 4, group1_col, models[0], selected_model)
        _write_model_group_headers(ws1, 4, group2_col, models[1], selected_model)

        for idx in range(fc_count):
            row = 5 + idx
            _write_model_group_data(ws1, row, group1_col, models[0], idx,
                                    forecast_data, py_forecasts, ind_forecasts, variance_data, frequency)
            _write_model_group_data(ws1, row, group2_col, models[1], idx,
                                    forecast_data, py_forecasts, ind_forecasts, variance_data, frequency)

        # Bottom two models: start after top models data + 2 blank rows
        bottom_header_row = 5 + fc_count + 2
        ws1.row_dimensions[bottom_header_row].height = 27.6
        _write_model_group_headers(ws1, bottom_header_row, group1_col, models[2], selected_model)
        _write_model_group_headers(ws1, bottom_header_row, group2_col, models[3], selected_model)

        for idx in range(fc_count):
            row = bottom_header_row + 1 + idx
            _write_model_group_data(ws1, row, group1_col, models[2], idx,
                                    forecast_data, py_forecasts, ind_forecasts, variance_data, frequency)
            _write_model_group_data(ws1, row, group2_col, models[3], idx,
                                    forecast_data, py_forecasts, ind_forecasts, variance_data, frequency)

    else:
        # ── Horizontal Layout (>12 predictions) ──
        spacers = [3, 10, 17, 24]
        group_starts = [4, 11, 18, 25]

        for s in spacers:
            _set_spacer(ws1, s)

        for i, model in enumerate(models):
            _write_model_group_headers(ws1, 4, group_starts[i], model, selected_model)
            _autofit_group(ws1, group_starts[i])

        for idx in range(fc_count):
            row = 5 + idx
            for i, model in enumerate(models):
                _write_model_group_data(ws1, row, group_starts[i], model, idx,
                                        forecast_data, py_forecasts, ind_forecasts, variance_data, frequency)

    # ── Sheet 2: Metrics Comparison ──────────────────────────────────
    ws2 = wb.create_sheet("Metrics Comparison")
    ws2.sheet_view.showGridLines = False
    ws2.sheet_view.zoomScale = 100

    _add_logo(ws2)

    ws2["A3"] = f"Market Pulse \u2014 Metrics Comparison ({freq_label})"
    ws2["A3"].font = title_font
    ws2["A3"].alignment = Alignment(horizontal="left", vertical="center")
    ws2.row_dimensions[3].height = 20.55

    # Helper functions for safe arithmetic
    def _safe_diff(a, b):
        """Variance = MP − R (preserves sign, positive = MP higher)."""
        if a is None or b is None:
            return None
        try:
            return round(float(a) - float(b), 2)
        except (TypeError, ValueError):
            return None

    def _safe_pct(a, b):
        """Variance % = (MP − R) / MP × 100. Returns None if MP ≈ 0 with real difference."""
        if a is None or b is None:
            return None
        try:
            fa, fb = float(a), float(b)
            if abs(fa) < 1e-6:
                return 0.0 if abs(fa - fb) < 1e-6 else None
            return round((fa - fb) / fa * 100, 2)
        except (TypeError, ValueError):
            return None

    def _safe_val(v):
        if v is None:
            return None
        try:
            f = float(v)
            return None if f == float("inf") or f == float("-inf") else f
        except (TypeError, ValueError):
            return None

    # 3 separate tables: MAE, SMAPE, MFE — each with own header row
    # 3-tier thresholds: (strong_limit, moderate_limit)
    # MAE/SMAPE: Strong ≤5%, Moderate 5–10%, Weak >10%
    # MFE: Strong ≤10%, Moderate 10–15%, Weak >15%
    metric_tables = [
        ("MAE", "mae", 5, 10),
        ("SMAPE", "smape", 5, 10),
        ("MFE", "mfe", 10, 15),
    ]

    table_headers_template = [
        ("Model", _fill_base),
        ("MP {metric}", _fill_mp),
        ("R {metric}", _fill_r),
        ("{metric} Variance", _fill_analysis),
        ("{metric} Variance %", _fill_analysis),
    ]

    row = 4
    for table_idx, (metric_label, metric_key, strong_limit, moderate_limit) in enumerate(metric_tables):
        # Header row
        ws2.row_dimensions[row].height = 27.6
        headers = [(h.format(metric=metric_label), f) for h, f in table_headers_template]
        for col_idx, (hdr, fill) in enumerate(headers, 1):
            _write_header_cell(ws2, row, col_idx, hdr, fill)
        # Left-align Column A header
        ws2.cell(row=row, column=1).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        row += 1

        # Data rows (one per model)
        for name in MODEL_ORDER:
            ind_m = ind_metrics.get(name, {"mae": 0, "smape": 0, "mfe": 0})
            py_m = py_metrics.get(name, {"mae": 0, "smape": 0, "mfe": 0})
            p_val = _safe_val(py_m.get(metric_key))
            i_val = _safe_val(ind_m.get(metric_key))
            var_diff = _safe_diff(p_val, i_val)
            var_pct = _safe_pct(p_val, i_val)

            vals = [name, p_val, i_val, var_diff, var_pct]

            for col_idx, val in enumerate(vals, 1):
                cell = ws2.cell(row=row, column=col_idx, value=val)
                cell.font = _data_font
                cell.alignment = Alignment(horizontal="left", vertical="center") if col_idx == 1 else _data_align
                cell.border = _border
                if col_idx == 5 and val is not None:
                    cell.number_format = '0.00"%"'
                elif col_idx in (2, 3) and metric_key == "smape" and isinstance(val, (int, float)):
                    cell.number_format = '0.00"%"'
                elif isinstance(val, float):
                    cell.number_format = _number_fmt

            # Color Variance + Variance %: 3-tier (green/yellow/red)
            if var_pct is not None:
                abs_pct = abs(var_pct)
                if abs_pct <= strong_limit:
                    fill, font = _green_fill, _green_font
                elif abs_pct <= moderate_limit:
                    fill, font = _yellow_fill, _yellow_font
                else:
                    fill, font = _var_highlight_fill, _var_highlight_font
                for c in (4, 5):
                    ws2.cell(row=row, column=c).fill = fill
                    ws2.cell(row=row, column=c).font = font

            row += 1

        # Blank row between tables (except after last)
        if table_idx < len(metric_tables) - 1:
            row += 1

    # Agreement Score — Column A label, Column B value, no merge, font size 12
    row += 1
    ws2.row_dimensions[row].height = 27.6
    _score_font_12 = Font(name="Calibri", bold=True, size=12, color="FFFFFF")

    label_cell = ws2.cell(row=row, column=1, value="Agreement Score")
    label_cell.font = _score_font_12
    label_cell.fill = _fill_base
    label_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    label_cell.border = _border

    if agreement_score is None:
        # Not Comparable — frequency mismatch
        _yellow_score_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
        score_cell = ws2.cell(row=row, column=2, value="Not Comparable")
        score_cell.font = Font(name="Calibri", bold=True, size=12, color="92400E")
        score_cell.fill = _yellow_score_fill
    elif agreement_score >= 90:
        score_cell = ws2.cell(row=row, column=2, value=agreement_score)
        score_cell.font = Font(name="Calibri", bold=True, size=12, color="065F46")
        score_cell.fill = _green_fill
        score_cell.number_format = "0.00"
    else:
        score_cell = ws2.cell(row=row, column=2, value=agreement_score)
        score_cell.font = Font(name="Calibri", bold=True, size=12, color="9C0006")
        score_cell.fill = _var_highlight_fill
        score_cell.number_format = "0.00"
    score_cell.alignment = _data_align
    score_cell.border = _border

    # Validation warnings/notes below Agreement Score
    _note_font = Font(name="Calibri", italic=True, size=9, color="6B7280")
    _note_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    for warning in (validation_warnings or []):
        row += 1
        ws2.row_dimensions[row].height = 30
        note_cell = ws2.cell(row=row, column=1, value=warning)
        note_cell.font = _note_font
        note_cell.alignment = _note_align
        ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)

    # Autofit columns — widest of header text or "Agreement Score"
    col_widths = [19, DATA_COL_WIDTH, DATA_COL_WIDTH, 18, 20]
    for col_idx, w in enumerate(col_widths, 1):
        ws2.column_dimensions[get_column_letter(col_idx)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
