"""Manual Validation Excel generator.

Sheet 1: Raw Data — uploaded dataset
Sheet 2: Model Parameters — ETS + ARIMA fitted parameters
Sheet 3: Your Forecasts — empty cells for user to fill (ETS full, ARIMA guided)
Sheet 4: Validation — auto-calculated variance via Excel formulas
"""

import io
from pathlib import Path

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.drawing.image import Image as XlImage
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from app.services.excel_export import LOGO_PATH, FREQ_LABELS, _format_date
from app.utils.logger import get_logger

logger = get_logger("manual_validation_export")

# Styling
_thin_side = Side(style="thin", color="E5E7EB")
_thin_border = Border(top=_thin_side, bottom=_thin_side, left=_thin_side, right=_thin_side)
_header_font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
_header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
_header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
_left_align = Alignment(horizontal="left", vertical="center")
_center_align = Alignment(horizontal="center", vertical="center")
_data_font = Font(name="Calibri", size=10, color="1F2937")
_number_fmt = "#,##0.00"
_yellow_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
_param_label_font = Font(name="Calibri", bold=True, size=10, color="1F2937")
_param_desc_font = Font(name="Calibri", italic=True, size=9, color="6B7280")
_instruction_font = Font(name="Calibri", italic=True, size=9, color="6B7280")
_worked_fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")  # blue tint for worked example


def _add_logo(ws):
    if LOGO_PATH.exists():
        logo = XlImage(str(LOGO_PATH))
        logo.width = 150
        logo.height = int(150 * 791 / 2160)
        ws.add_image(logo, "A1")
    ws.row_dimensions[1].height = 20.5
    ws.row_dimensions[2].height = 20.5


def generate_manual_validation_excel(
    historical_data: list[dict],
    forecast_data: list[dict],
    comparison_forecasts: dict[str, list[float]],
    frequency: str,
    model_params: dict,
) -> bytes:
    """Generate the Manual Validation Excel workbook."""
    freq_label = FREQ_LABELS.get(frequency, frequency)
    wb = Workbook()
    title_font = Font(name="Calibri", bold=True, size=14, color="1F2937")

    ets_params = model_params.get("ets", {})
    arima_params = model_params.get("arima", {})
    fc_count = len(forecast_data)

    # ── Sheet 1: Raw Data ────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Raw Data"
    ws1.sheet_view.showGridLines = False

    _add_logo(ws1)

    ws1["A3"] = f"Market Pulse — Raw Data ({freq_label})"
    ws1["A3"].font = title_font
    ws1["A3"].alignment = _left_align
    ws1.row_dimensions[3].height = 20.5

    # Headers (row 4 — no blank row 4)
    _left_header_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    for col_idx, hdr in enumerate(["Date", "Value"], 1):
        cell = ws1.cell(row=4, column=col_idx, value=hdr)
        cell.font = _header_font
        cell.fill = _header_fill
        cell.alignment = _left_header_align if col_idx == 1 else _header_align
        cell.border = _thin_border

    row = 5
    for entry in historical_data:
        date_str = _format_date(entry["date"], frequency)
        ws1.cell(row=row, column=1, value=date_str).font = _data_font
        ws1.cell(row=row, column=1).alignment = _left_align
        ws1.cell(row=row, column=1).border = _thin_border
        val_cell = ws1.cell(row=row, column=2, value=entry["value"])
        val_cell.font = _data_font
        val_cell.number_format = _number_fmt
        val_cell.alignment = _center_align
        val_cell.border = _thin_border
        row += 1

    ws1.column_dimensions["A"].width = 16
    ws1.column_dimensions["B"].width = 16

    # ── Sheet 2: Model Parameters ────────────────────────────────────
    ws2 = wb.create_sheet("Model Parameters")
    ws2.sheet_view.showGridLines = False

    _add_logo(ws2)

    ws2["A3"] = "Market Pulse — Model Parameters"
    ws2["A3"].font = title_font
    ws2["A3"].alignment = _left_align
    ws2.row_dimensions[3].height = 20.5

    # ETS Section (row 4 — no blank row)
    row = 4
    ws2.cell(row=row, column=1, value="ETS Parameters").font = Font(name="Calibri", bold=True, size=12, color="1F2937")
    row += 1

    ets_fields = [
        ("Alpha (α)", ets_params.get("alpha"), "Smoothing level — weight given to recent observations"),
        ("Beta (β)", ets_params.get("beta"), "Smoothing trend — weight for trend component"),
        ("Gamma (γ)", ets_params.get("gamma"), "Smoothing seasonal — weight for seasonal component"),
        ("Initial Level (l₀)", ets_params.get("l0"), "Starting level value"),
        ("Initial Trend (b₀)", ets_params.get("b0"), "Starting trend value"),
        ("Trend Type", ets_params.get("trend", "None"), "Additive, multiplicative, or none"),
        ("Seasonal Type", ets_params.get("seasonal", "None"), "Additive, multiplicative, or none"),
        ("Seasonal Period", ets_params.get("seasonal_period", "None"), "Number of periods in one seasonal cycle"),
    ]

    for label, value, desc in ets_fields:
        ws2.cell(row=row, column=1, value=label).font = _param_label_font
        ws2.cell(row=row, column=1).border = _thin_border
        val_cell = ws2.cell(row=row, column=2, value=value if value is not None else "N/A")
        val_cell.font = _data_font
        val_cell.alignment = _center_align
        val_cell.border = _thin_border
        if isinstance(value, float):
            val_cell.number_format = "0.000000"
        ws2.cell(row=row, column=3, value=desc).font = _param_desc_font
        ws2.cell(row=row, column=3).border = _thin_border
        row += 1

    # ARIMA Section
    row += 1
    ws2.cell(row=row, column=1, value="ARIMA Parameters").font = Font(name="Calibri", bold=True, size=12, color="1F2937")
    row += 1

    order = arima_params.get("order", (0, 0, 0))
    seasonal_order = arima_params.get("seasonal_order", (0, 0, 0, 1))

    # Order (p,d,q) row
    ws2.cell(row=row, column=1, value="Order (p,d,q)").font = _param_label_font
    ws2.cell(row=row, column=1).border = _thin_border
    ws2.cell(row=row, column=2, value=f"({order[0]}, {order[1]}, {order[2]})").font = _data_font
    ws2.cell(row=row, column=2).alignment = _center_align
    ws2.cell(row=row, column=2).border = _thin_border
    ws2.cell(row=row, column=3, value="AR order, differencing, MA order").font = _param_desc_font
    ws2.cell(row=row, column=3).border = _thin_border
    row += 1

    # d row — differencing order extracted separately
    ws2.cell(row=row, column=1, value="d").font = _param_label_font
    ws2.cell(row=row, column=1).border = _thin_border
    ws2.cell(row=row, column=2, value=order[1]).font = _data_font
    ws2.cell(row=row, column=2).alignment = _center_align
    ws2.cell(row=row, column=2).border = _thin_border
    ws2.cell(row=row, column=3, value="Differencing order — how many times the data is differenced to remove trend").font = _param_desc_font
    ws2.cell(row=row, column=3).border = _thin_border
    row += 1

    # Seasonal Order row
    ws2.cell(row=row, column=1, value="Seasonal Order (P,D,Q,m)").font = _param_label_font
    ws2.cell(row=row, column=1).border = _thin_border
    ws2.cell(row=row, column=2, value=f"({seasonal_order[0]}, {seasonal_order[1]}, {seasonal_order[2]}, {seasonal_order[3]})").font = _data_font
    ws2.cell(row=row, column=2).alignment = _center_align
    ws2.cell(row=row, column=2).border = _thin_border
    ws2.cell(row=row, column=3, value="Seasonal AR, differencing, MA, period").font = _param_desc_font
    ws2.cell(row=row, column=3).border = _thin_border
    row += 1

    # Coefficients
    coefficients = arima_params.get("coefficients", {})
    if coefficients:
        row += 1
        ws2.cell(row=row, column=1, value="Fitted Coefficients").font = Font(name="Calibri", bold=True, size=11, color="1F2937")
        row += 1
        for name, val in coefficients.items():
            ws2.cell(row=row, column=1, value=name).font = _param_label_font
            ws2.cell(row=row, column=1).border = _thin_border
            val_cell = ws2.cell(row=row, column=2, value=val)
            val_cell.font = _data_font
            val_cell.alignment = _center_align
            val_cell.number_format = "0.000000"
            val_cell.border = _thin_border
            row += 1

    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 20
    # Autofit column C based on longest description
    max_c_len = 0
    for r in range(4, ws2.max_row + 1):
        val = ws2.cell(row=r, column=3).value
        if val:
            max_c_len = max(max_c_len, len(str(val)))
    ws2.column_dimensions["C"].width = min(max_c_len + 2, 80)  # cap at 80

    # ── Sheet 3: Your Forecasts ──────────────────────────────────────
    ws3 = wb.create_sheet("Your Forecasts")
    ws3.sheet_view.showGridLines = False

    _add_logo(ws3)

    ws3["A3"] = f"Market Pulse — Your Forecasts ({freq_label})"
    ws3["A3"].font = title_font
    ws3["A3"].alignment = _left_align
    ws3.row_dimensions[3].height = 20.5

    # Headers (row 4)
    _left_header = Alignment(horizontal="left", vertical="center", wrap_text=True)
    yf_headers = ["Date", "ETS Forecast (FILL)", "ARIMA Forecast (FILL)", "Instructions"]
    for col_idx, hdr in enumerate(yf_headers, 1):
        cell = ws3.cell(row=4, column=col_idx, value=hdr)
        cell.font = _header_font
        cell.fill = _header_fill
        cell.alignment = _left_header if col_idx in (1, 4) else _header_align
        cell.border = _thin_border

    # --- Dynamic formula computation ---
    # Extract last historical values for worked examples
    hist_values = [h["value"] for h in historical_data]
    last_val = hist_values[-1] if hist_values else 0
    prev_val = hist_values[-2] if len(hist_values) >= 2 else last_val

    # ETS parameters
    alpha = ets_params.get("alpha", 1)
    trend_type = ets_params.get("trend", None)
    seasonal_type = ets_params.get("seasonal", None)

    # ARIMA parameters
    d = order[1]
    p, q = order[0], order[2]
    coefficients = arima_params.get("coefficients", {})
    intercept = coefficients.get("intercept", 0)
    ar1 = coefficients.get("ar.L1", 0)
    ma1 = coefficients.get("ma.L1", 0)
    intercept_display = round(intercept, 2) if abs(intercept) >= 0.1 else round(intercept, 4)

    # Get our predictions for worked examples
    ets_preds = comparison_forecasts.get("AutoETS", [])
    arima_preds = comparison_forecasts.get("AutoARIMA", [])

    # Dynamic worked example count: d=2 gets 5, all others get 3
    N_WORKED = 5 if d == 2 else 3

    # --- Compute ETS worked examples ---
    def _compute_ets_examples(n=3):
        if trend_type is None and seasonal_type is None:
            return [round(last_val, 2)] * n
        elif trend_type == "add" and seasonal_type is None:
            trend_val = round(last_val - prev_val, 2)
            return [round(last_val + (h + 1) * trend_val, 2) for h in range(n)]
        else:
            # For seasonal or complex models, use model's own predictions
            return [round(ets_preds[h], 2) if h < len(ets_preds) else None for h in range(n)]

    ets_examples = _compute_ets_examples(N_WORKED)

    # --- Compute ARIMA worked examples (first 3 rows) ---
    def _compute_arima_examples(n=3):
        vals = []
        if d == 0 and p >= 1:
            # AR model: Next = intercept + ar1 × previous
            prev = last_val
            for i in range(n):
                if i == 0 and arima_preds:
                    vals.append(round(arima_preds[0], 2))
                    prev = arima_preds[0]
                else:
                    nxt = intercept + ar1 * prev
                    vals.append(round(nxt, 2))
                    prev = nxt
        elif d == 1:
            # d=1: Next = prev + intercept (simplified, AR/MA terms decay)
            prev = last_val
            for i in range(n):
                if i == 0 and arima_preds:
                    vals.append(round(arima_preds[0], 2))
                    prev = arima_preds[0]
                else:
                    nxt = prev + intercept
                    vals.append(round(nxt, 2))
                    prev = nxt
        elif d == 2:
            # d=2: Next = 2×prev - prev2 + intercept
            prev2, prev1 = prev_val, last_val
            for i in range(n):
                if i == 0 and arima_preds:
                    vals.append(round(arima_preds[0], 2))
                    prev2, prev1 = prev1, arima_preds[0]
                else:
                    nxt = 2 * prev1 - prev2 + intercept
                    vals.append(round(nxt, 2))
                    prev2, prev1 = prev1, nxt
        else:
            # d=0 without AR — use model predictions
            for i in range(n):
                if i < len(arima_preds):
                    vals.append(round(arima_preds[i], 2))
                else:
                    vals.append(None)
        return vals

    arima_examples = _compute_arima_examples(N_WORKED)

    # --- Dynamic ETS header comment ---
    if trend_type is None and seasonal_type is None:
        ets_comment_text = f"ETS Forecast = {round(last_val, 2)} (last observed value, repeated for all periods)"
    elif trend_type == "add" and seasonal_type is None:
        trend_val = round(last_val - prev_val, 2)
        ets_comment_text = f"ETS Forecast = Level + [months ahead \u00d7 Trend]\nLevel = {round(last_val, 2)}, Trend = {trend_val}"
    elif seasonal_type is not None:
        ets_comment_text = "ETS Forecast = Level + h \u00d7 Trend + Seasonal[h]\nSee Model Parameters for values"
    else:
        ets_comment_text = f"ETS Forecast = Level + [months ahead \u00d7 Trend]\nLevel = {round(last_val, 2)}"

    b4_comment = Comment(ets_comment_text, "Market Pulse")
    b4_comment.width = 350
    b4_comment.height = 80
    ws3.cell(row=4, column=2).comment = b4_comment

    # --- Dynamic ARIMA header comment (always show d-based formula) ---
    # Build coefficient summary for complex models
    coeff_parts = []
    for cname, cval in coefficients.items():
        if cname not in ("intercept", "sigma2"):
            coeff_parts.append(f"{cname} = {round(cval, 4)}")
    coeff_summary = ", ".join(coeff_parts) if coeff_parts else ""

    if d == 0 and p >= 1 and q >= 1:
        arima_comment_text = (
            f"ARIMA Forecast:\n"
            f"Row 1: {intercept_display} + {round(ar1, 4)} \u00d7 last_value + {round(ma1, 4)} \u00d7 residual\n"
            f"Row 2+: {intercept_display} + {round(ar1, 4)} \u00d7 previous"
        )
    elif d == 0 and p >= 1:
        arima_comment_text = f"ARIMA Forecast = {intercept_display} + {round(ar1, 4)} \u00d7 previous value"
    elif d == 1:
        arima_comment_text = f"ARIMA Forecast (d=1) = previous + intercept of {intercept_display}"
        if coeff_summary:
            arima_comment_text += f"\nCoefficients: {coeff_summary}\nSee Model Parameters for full formula"
    elif d == 2:
        arima_comment_text = f"ARIMA Forecast (d=2) = 2 \u00d7 previous \u2212 one_before + intercept of {intercept_display}"
        if coeff_summary:
            arima_comment_text += f"\nCoefficients: {coeff_summary}\nSee Model Parameters for full formula"
    else:
        arima_comment_text = f"ARIMA({p},{d},{q}) \u2014 intercept = {intercept_display}"
        if coeff_summary:
            arima_comment_text += f"\nCoefficients: {coeff_summary}"

    c4_comment = Comment(arima_comment_text, "Market Pulse")
    c4_comment.width = 400
    c4_comment.height = 100
    ws3.cell(row=4, column=3).comment = c4_comment

    # --- Build per-row instructions ---
    def _ets_instruction(idx):
        if trend_type is None and seasonal_type is None:
            return f"= {round(last_val, 2)} (same for all rows)"
        elif trend_type == "add" and seasonal_type is None:
            trend_val = round(last_val - prev_val, 2)
            h = idx + 1
            result = round(last_val + h * trend_val, 2)
            return f"= level ({round(last_val, 2)}) + {h} \u00d7 trend ({trend_val}) = {result}"
        elif seasonal_type is not None:
            if idx < len(ets_preds):
                return f"= {round(ets_preds[idx], 2)} (uses seasonal pattern from Model Parameters)"
            return "Continue seasonal pattern"
        else:
            return f"= {round(last_val, 2)}"

    def _arima_instruction(idx):
        if idx < len(arima_examples) and arima_examples[idx] is not None:
            val = arima_examples[idx]
            prev = last_val if idx == 0 else arima_examples[idx - 1]
            if d == 0 and p >= 1:
                if idx == 0:
                    return f"= intercept ({intercept_display}) + ar1 ({round(ar1, 4)}) \u00d7 last ({round(last_val, 2)}) + MA term = {val}"
                return f"= intercept ({intercept_display}) + ar1 ({round(ar1, 4)}) \u00d7 previous ({round(prev, 2)}) = {val}"
            elif d == 1:
                if idx == 0 and (p > 0 or q > 0):
                    return f"= {val} (Row 1 includes AR/MA terms \u2014 verify with Model Parameters)"
                return f"= previous ({round(prev, 2)}) + intercept ({intercept_display}) = {val}"
            elif d == 2:
                prev2 = prev_val if idx == 0 else (last_val if idx == 1 else arima_examples[idx - 2])
                return f"= 2 \u00d7 previous ({round(prev, 2)}) \u2212 one_before ({round(prev2, 2)}) + intercept ({intercept_display}) = {val}"
            else:
                return f"= {val} (see Model Parameters for formula)"
        if d == 1:
            return f"= previous + {intercept_display}"
        elif d == 2:
            return f"= 2 \u00d7 previous \u2212 one_before + {intercept_display}"
        return "Continue the pattern \u2014 each row uses the same formula"

    row = 5
    for idx, entry in enumerate(forecast_data):
        date_str = _format_date(entry["date"], frequency)
        ws3.cell(row=row, column=1, value=date_str).font = _data_font
        ws3.cell(row=row, column=1).alignment = _left_align
        ws3.cell(row=row, column=1).border = _thin_border

        # ETS column: first 3 rows = blue worked examples, rest = yellow
        if idx < N_WORKED and idx < len(ets_examples) and ets_examples[idx] is not None:
            ets_cell = ws3.cell(row=row, column=2, value=ets_examples[idx])
            ets_cell.fill = _worked_fill
            ets_cell.font = _data_font
        else:
            ets_cell = ws3.cell(row=row, column=2, value="")
            ets_cell.fill = _yellow_fill
        ets_cell.alignment = _center_align
        ets_cell.border = _thin_border
        ets_cell.number_format = _number_fmt

        # ARIMA column: first 3 rows = blue worked examples, rest = yellow
        if idx < N_WORKED and idx < len(arima_examples) and arima_examples[idx] is not None:
            arima_cell = ws3.cell(row=row, column=3, value=arima_examples[idx])
            arima_cell.fill = _worked_fill
            arima_cell.font = _data_font
        else:
            arima_cell = ws3.cell(row=row, column=3, value="")
            arima_cell.fill = _yellow_fill
        arima_cell.alignment = _center_align
        arima_cell.border = _thin_border
        arima_cell.number_format = _number_fmt

        # Instructions column — dynamic per-row
        if idx < N_WORKED:
            instruction = f"ETS: {_ets_instruction(idx)}  |  ARIMA: {_arima_instruction(idx)}"
        else:
            instruction = "Continue the pattern \u2014 each row uses the same formula"

        instr_cell = ws3.cell(row=row, column=4, value=instruction)
        instr_cell.font = _instruction_font
        instr_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        instr_cell.border = _thin_border

        row += 1

    ws3.column_dimensions["A"].width = 16
    ws3.column_dimensions["B"].width = 22
    ws3.column_dimensions["C"].width = 24
    # Autofit column D based on longest instruction value
    max_d_len = len("Instructions")  # header length as minimum
    for r in range(5, 5 + fc_count):
        val = ws3.cell(row=r, column=4).value
        if val:
            max_d_len = max(max_d_len, len(str(val)))
    ws3.column_dimensions["D"].width = min(max_d_len + 2, 80)  # cap at 80

    # ── Sheet 4: Validation ──────────────────────────────────────────
    ws4 = wb.create_sheet("Validation")
    ws4.sheet_view.showGridLines = False

    _add_logo(ws4)

    ws4["A3"] = f"Market Pulse — Validation ({freq_label})"
    ws4["A3"].font = title_font
    ws4["A3"].alignment = _left_align
    ws4.row_dimensions[3].height = 20.5

    # Two side-by-side tables with spacer column F
    # ETS table: A-E, Spacer: F, ARIMA table: G-K

    # ETS Headers (row 4)
    _var_header_fill = PatternFill(start_color="495057", end_color="495057", fill_type="solid")
    _manual_header_fill = PatternFill(start_color="C3791F", end_color="C3791F", fill_type="solid")
    _model_header_fill = PatternFill(start_color="248479", end_color="248479", fill_type="solid")
    _white_header_font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")

    # ETS: col1=Date(default), col2=Manual(orange), col3=Model(teal), col4-5=Variance(gray)
    ets_headers = ["Date", "Manual ETS (Calculated)", "Model ETS (Predicted)", "Variance (ETS)", "Variance % (ETS)"]
    ets_fills = {1: _header_fill, 2: _manual_header_fill, 3: _model_header_fill, 4: _var_header_fill, 5: _var_header_fill}
    for col_idx, hdr in enumerate(ets_headers, 1):
        cell = ws4.cell(row=4, column=col_idx, value=hdr)
        cell.font = _white_header_font
        cell.fill = ets_fills[col_idx]
        cell.alignment = _left_align if col_idx == 1 else _header_align
        cell.border = _thin_border

    # ARIMA: col7=Date(default), col8=Manual(orange), col9=Model(teal), col10-11=Variance(gray)
    arima_headers = ["Date", "Manual ARIMA (Calculated)", "Model ARIMA (Predicted)", "Variance (ARIMA)", "Variance % (ARIMA)"]
    arima_fills = {7: _header_fill, 8: _manual_header_fill, 9: _model_header_fill, 10: _var_header_fill, 11: _var_header_fill}
    for col_idx, hdr in enumerate(arima_headers, 7):
        cell = ws4.cell(row=4, column=col_idx, value=hdr)
        cell.font = _white_header_font
        cell.fill = arima_fills[col_idx]
        cell.alignment = _left_align if col_idx == 7 else _header_align
        cell.border = _thin_border

    # Sheet 3 data starts at row 5 now
    sheet3_data_start = 5

    row = 5
    for idx, entry in enumerate(forecast_data):
        date_str = _format_date(entry["date"], frequency)
        sheet3_row = sheet3_data_start + idx

        # ── ETS side (A-E) ──
        # A: Date
        ws4.cell(row=row, column=1, value=date_str).font = _data_font
        ws4.cell(row=row, column=1).alignment = _left_align
        ws4.cell(row=row, column=1).border = _thin_border

        # B: Manual ETS = reference to Sheet 3 column B
        ets_ref_cell = ws4.cell(row=row, column=2)
        ets_ref_cell.value = f"='Your Forecasts'!B{sheet3_row}"
        ets_ref_cell.font = _data_font
        ets_ref_cell.number_format = _number_fmt
        ets_ref_cell.alignment = _center_align
        ets_ref_cell.border = _thin_border

        # C: Model ETS
        our_ets = ets_preds[idx] if idx < len(ets_preds) else None
        our_ets_cell = ws4.cell(row=row, column=3, value=our_ets)
        our_ets_cell.font = _data_font
        our_ets_cell.number_format = _number_fmt
        our_ets_cell.alignment = _center_align
        our_ets_cell.border = _thin_border

        # D: Variance ETS = Manual - Model
        var_ets_cell = ws4.cell(row=row, column=4)
        var_ets_cell.value = f"=B{row}-C{row}"
        var_ets_cell.font = _data_font
        var_ets_cell.number_format = _number_fmt
        var_ets_cell.alignment = _center_align
        var_ets_cell.border = _thin_border

        # E: Variance % ETS
        var_pct_ets_cell = ws4.cell(row=row, column=5)
        var_pct_ets_cell.value = f"=IF(C{row}=0,0,(B{row}-C{row})/C{row}*100)"
        var_pct_ets_cell.font = _data_font
        var_pct_ets_cell.number_format = '0.00"%"'
        var_pct_ets_cell.alignment = _center_align
        var_pct_ets_cell.border = _thin_border

        # F: Spacer (empty)

        # ── ARIMA side (G-K) ──
        # G: Date
        ws4.cell(row=row, column=7, value=date_str).font = _data_font
        ws4.cell(row=row, column=7).alignment = _left_align
        ws4.cell(row=row, column=7).border = _thin_border

        # H: Manual ARIMA = reference to Sheet 3 column C
        arima_ref_cell = ws4.cell(row=row, column=8)
        arima_ref_cell.value = f"='Your Forecasts'!C{sheet3_row}"
        arima_ref_cell.font = _data_font
        arima_ref_cell.number_format = _number_fmt
        arima_ref_cell.alignment = _center_align
        arima_ref_cell.border = _thin_border

        # I: Model ARIMA
        our_arima = arima_preds[idx] if idx < len(arima_preds) else None
        our_arima_cell = ws4.cell(row=row, column=9, value=our_arima)
        our_arima_cell.font = _data_font
        our_arima_cell.number_format = _number_fmt
        our_arima_cell.alignment = _center_align
        our_arima_cell.border = _thin_border

        # J: Variance ARIMA = Manual - Model
        var_arima_cell = ws4.cell(row=row, column=10)
        var_arima_cell.value = f"=H{row}-I{row}"
        var_arima_cell.font = _data_font
        var_arima_cell.number_format = _number_fmt
        var_arima_cell.alignment = _center_align
        var_arima_cell.border = _thin_border

        # K: Variance % ARIMA
        var_pct_arima_cell = ws4.cell(row=row, column=11)
        var_pct_arima_cell.value = f"=IF(I{row}=0,0,(H{row}-I{row})/I{row}*100)"
        var_pct_arima_cell.font = _data_font
        var_pct_arima_cell.number_format = '0.00"%"'
        var_pct_arima_cell.alignment = _center_align
        var_pct_arima_cell.border = _thin_border

        row += 1

    # Conditional formatting for variance columns (D, E, J, K)
    # 3-tier: Strong (green) ≤1%, Moderate (yellow) 1–5%, Weak (red) >5%
    _cf_green_fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5")
    _cf_green_font = Font(name="Calibri", bold=True, size=10, color="065F46")
    _cf_yellow_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7")
    _cf_yellow_font = Font(name="Calibri", bold=True, size=10, color="92400E")
    _cf_red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE")
    _cf_red_font = Font(name="Calibri", size=10, color="9C0006")
    fc_start = 5
    fc_end = 5 + fc_count - 1

    # ETS side: color D & E based on E (Variance %) — order: red first, yellow, green last
    for col_letter in ["D", "E"]:
        rng = f"{col_letter}{fc_start}:{col_letter}{fc_end}"
        ws4.conditional_formatting.add(rng, FormulaRule(
            formula=[f"ABS($E{fc_start})>5"],
            fill=_cf_red_fill, font=_cf_red_font,
        ))
        ws4.conditional_formatting.add(rng, FormulaRule(
            formula=[f"ABS($E{fc_start})>1"],
            fill=_cf_yellow_fill, font=_cf_yellow_font,
        ))
        ws4.conditional_formatting.add(rng, FormulaRule(
            formula=[f"ABS($E{fc_start})<=1"],
            fill=_cf_green_fill, font=_cf_green_font,
        ))

    # ARIMA side: color J & K based on K (Variance %) — order: red first, yellow, green last
    for col_letter in ["J", "K"]:
        rng = f"{col_letter}{fc_start}:{col_letter}{fc_end}"
        ws4.conditional_formatting.add(rng, FormulaRule(
            formula=[f"ABS($K{fc_start})>5"],
            fill=_cf_red_fill, font=_cf_red_font,
        ))
        ws4.conditional_formatting.add(rng, FormulaRule(
            formula=[f"ABS($K{fc_start})>1"],
            fill=_cf_yellow_fill, font=_cf_yellow_font,
        ))
        ws4.conditional_formatting.add(rng, FormulaRule(
            formula=[f"ABS($K{fc_start})<=1"],
            fill=_cf_green_fill, font=_cf_green_font,
        ))

    # MAE Summary rows
    row += 1

    # MAE row styling — matches Agreement Score from Independent Validation
    _mae_label_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    _mae_label_font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
    ws4.row_dimensions[row].height = 32

    # ETS MAE (cols A-B)
    ws4.cell(row=row, column=1, value="MAE (Manual Calculated ETS)").font = _mae_label_font
    ws4.cell(row=row, column=1).fill = _mae_label_fill
    ws4.cell(row=row, column=1).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws4.cell(row=row, column=1).border = _thin_border
    mae_ets_cell = ws4.cell(row=row, column=2)
    mae_ets_cell.value = f"=SUMPRODUCT(ABS(D{fc_start}:D{fc_end}))/COUNT(D{fc_start}:D{fc_end})"
    mae_ets_cell.font = Font(name="Calibri", bold=True, size=12)
    mae_ets_cell.number_format = "0.00"
    mae_ets_cell.alignment = _center_align
    mae_ets_cell.border = _thin_border

    # ARIMA MAE (cols G-H)
    ws4.cell(row=row, column=7, value="MAE (Manual Calculated ARIMA)").font = _mae_label_font
    ws4.cell(row=row, column=7).fill = _mae_label_fill
    ws4.cell(row=row, column=7).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws4.cell(row=row, column=7).border = _thin_border
    mae_arima_cell = ws4.cell(row=row, column=8)
    mae_arima_cell.value = f"=SUMPRODUCT(ABS(J{fc_start}:J{fc_end}))/COUNT(J{fc_start}:J{fc_end})"
    mae_arima_cell.font = Font(name="Calibri", bold=True, size=12)
    mae_arima_cell.number_format = "0.00"
    mae_arima_cell.alignment = _center_align
    mae_arima_cell.border = _thin_border

    # Dynamic 3-tier MAE color coding: normalize by avg forecast magnitude
    # MAE % = MAE / avg(|predictions|) × 100
    # Green ≤ 1%, Yellow 1-5%, Red > 5%
    _mae_green_font = Font(name="Calibri", bold=True, size=12, color="065F46")
    _mae_yellow_font = Font(name="Calibri", bold=True, size=12, color="92400E")
    _mae_red_font = Font(name="Calibri", bold=True, size=12, color="9C0006")
    _cf_yellow_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")

    avg_ets = round(sum(abs(v) for v in ets_preds if v) / max(len(ets_preds), 1), 6)
    avg_arima = round(sum(abs(v) for v in arima_preds if v) / max(len(arima_preds), 1), 6)

    for mae_col, avg_val in [(f"B{row}", avg_ets), (f"H{row}", avg_arima)]:
        if avg_val > 0:
            # Green: MAE % ≤ 1%  →  MAE ≤ avg * 0.01
            t1 = round(avg_val * 0.01, 6)
            # Yellow: MAE % ≤ 5%  →  MAE ≤ avg * 0.05
            t5 = round(avg_val * 0.05, 6)
            ws4.conditional_formatting.add(mae_col, CellIsRule(
                operator="lessThanOrEqual", formula=[str(t1)],
                fill=_cf_green_fill, font=_mae_green_font,
            ))
            ws4.conditional_formatting.add(mae_col, FormulaRule(
                formula=[f"AND({mae_col}>{t1},{mae_col}<={t5})"],
                fill=_cf_yellow_fill, font=_mae_yellow_font,
            ))
            ws4.conditional_formatting.add(mae_col, CellIsRule(
                operator="greaterThan", formula=[str(t5)],
                fill=_cf_red_fill, font=_mae_red_font,
            ))
        else:
            # All predictions are 0 — any MAE is green
            ws4.conditional_formatting.add(mae_col, CellIsRule(
                operator="greaterThanOrEqual", formula=["0"],
                fill=_cf_green_fill, font=_mae_green_font,
            ))

    # Column widths
    ws4.column_dimensions["A"].width = 17  # Date
    for col in ["B", "C", "D", "E"]:
        ws4.column_dimensions[col].width = 13  # value columns
    ws4.column_dimensions["F"].width = 5  # spacer
    ws4.column_dimensions["G"].width = 22  # Date
    for col in ["H", "I", "J", "K"]:
        ws4.column_dimensions[col].width = 13  # value columns

    # Write to bytes
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
