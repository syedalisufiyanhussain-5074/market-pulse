import io
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XlImage
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from app.utils.logger import get_logger, log_stage

LOGO_PATH = Path(__file__).parent.parent / "assets" / "logo-bold-horizontal.png"

logger = get_logger("excel_export")

# Display name mapping
DISPLAY_NAMES = {
    "AutoETS": "ETS",
    "AutoARIMA": "ARIMA",
}

# Frequency display names
FREQ_LABELS = {
    "D": "Daily",
    "W": "Weekly",
    "MS": "Monthly",
    "M": "Monthly",
    "QS": "Quarterly",
    "Q": "Quarterly",
    "YS": "Yearly",
    "Y": "Yearly",
}

# Date format per frequency
DATE_FORMATS = {
    "D": "%Y-%m-%d",
    "W": "%Y-%m-%d",
    "MS": "%b %Y",
    "M": "%b %Y",
    "QS": "Q%q %Y",
    "Q": "Q%q %Y",
    "YS": "%Y",
    "Y": "%Y",
}


def _format_date(iso_date: str, frequency: str) -> str:
    """Format an ISO date string according to frequency."""
    dt = datetime.fromisoformat(iso_date)
    fmt = DATE_FORMATS.get(frequency, "%Y-%m-%d")
    if "%q" in fmt:
        quarter = (dt.month - 1) // 3 + 1
        return fmt.replace("%q", str(quarter)).replace("%Y", str(dt.year))
    return dt.strftime(fmt)


def generate_excel(
    selected_model: str,
    historical_data: list[dict],
    forecast_data: list[dict],
    frequency: str,
    forecast_bias: str = "Forecast",
    comparison_forecasts: dict[str, list[float]] | None = None,
) -> bytes:
    """Generate a clean Excel workbook with historical + forecast data."""
    model_display = DISPLAY_NAMES.get(selected_model, selected_model)
    freq_label = FREQ_LABELS.get(frequency, frequency)

    wb = Workbook()
    ws = wb.active
    ws.title = "Forecast Data"

    # Styles
    header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )
    data_font = Font(name="Calibri", size=11)
    data_align = Alignment(horizontal="center", vertical="center")
    number_format = "#,##0.00"

    # Remove default gridlines
    ws.sheet_view.showGridLines = False

    # Logo (rows 1-2)
    if LOGO_PATH.exists():
        logo = XlImage(str(LOGO_PATH))
        logo.width = 150
        logo.height = int(150 * 791 / 2160)  # maintain aspect ratio
        ws.add_image(logo, "A1")
    ws.row_dimensions[1].height = 20.5
    ws.row_dimensions[2].height = 20.5

    # Title row (row 3) with same border as table
    title_font = Font(name="Calibri", bold=True, size=14, color="1F2937")
    ws.merge_cells("A3:E3")
    ws["A3"] = f"Market Pulse — {model_display} {forecast_bias} ({freq_label})"
    ws["A3"].font = title_font
    ws["A3"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[3].height = 20.5
    for col_idx in range(1, 6):
        ws.cell(row=3, column=col_idx).border = thin_border

    # Headers (row 5)
    headers = ["Date", "Actual", "Forecast", "Lower Bound", "Upper Bound"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Historical data
    row = 6
    for entry in historical_data:
        date_str = _format_date(entry["date"], frequency)
        ws.cell(row=row, column=1, value=date_str).font = data_font
        ws.cell(row=row, column=1).alignment = data_align
        ws.cell(row=row, column=1).border = thin_border

        val_cell = ws.cell(row=row, column=2, value=entry["value"])
        val_cell.font = data_font
        val_cell.number_format = number_format
        val_cell.alignment = data_align
        val_cell.border = thin_border

        # Forecast columns blank for historical
        for col in [3, 4, 5]:
            c = ws.cell(row=row, column=col, value="")
            c.border = thin_border
        row += 1

    # Forecast data
    for entry in forecast_data:
        date_str = _format_date(entry["date"], frequency)
        ws.cell(row=row, column=1, value=date_str).font = data_font
        ws.cell(row=row, column=1).alignment = data_align
        ws.cell(row=row, column=1).border = thin_border

        # Actual column blank for forecast
        c = ws.cell(row=row, column=2, value="")
        c.border = thin_border

        fc_cell = ws.cell(row=row, column=3, value=entry["value"])
        fc_cell.font = data_font
        fc_cell.number_format = number_format
        fc_cell.alignment = data_align
        fc_cell.border = thin_border

        if "lower_bound" in entry:
            lo_cell = ws.cell(row=row, column=4, value=entry["lower_bound"])
            lo_cell.font = data_font
            lo_cell.number_format = number_format
            lo_cell.alignment = data_align
            lo_cell.border = thin_border
        else:
            ws.cell(row=row, column=4, value="").border = thin_border

        if "upper_bound" in entry:
            hi_cell = ws.cell(row=row, column=5, value=entry["upper_bound"])
            hi_cell.font = data_font
            hi_cell.number_format = number_format
            hi_cell.alignment = data_align
            hi_cell.border = thin_border
        else:
            ws.cell(row=row, column=5, value="").border = thin_border

        row += 1

    # Auto-fit column widths (account for number formatting)
    for col_idx in range(1, 6):
        col_letter = get_column_letter(col_idx)
        max_len = len(headers[col_idx - 1])
        for r in range(6, row):
            cell = ws.cell(row=r, column=col_idx)
            val = cell.value
            if val is None or val == "":
                continue
            if isinstance(val, (int, float)):
                # Format as displayed: #,##0.00
                display_str = f"{val:,.2f}"
            else:
                display_str = str(val)
            max_len = max(max_len, len(display_str))
        ws.column_dimensions[col_letter].width = max_len + 4

    # ── Tab 2: Model Comparison ──────────────────────────────────────
    if comparison_forecasts:
        MODEL_ORDER = ["AutoETS", "AutoARIMA", "Moving Average (Excel)", "ETS (Excel)"]
        ws2 = wb.create_sheet("Model Comparison")
        ws2.sheet_view.showGridLines = False

        # Logo (rows 1-2)
        if LOGO_PATH.exists():
            logo2 = XlImage(str(LOGO_PATH))
            logo2.width = 150
            logo2.height = int(150 * 791 / 2160)
            ws2.add_image(logo2, "A1")
        ws2.row_dimensions[1].height = 20.5
        ws2.row_dimensions[2].height = 20.5

        # Title row (row 3)
        title_font2 = Font(name="Calibri", bold=True, size=14, color="1F2937")
        ws2.merge_cells("A3:F3")
        ws2["A3"] = f"Market Pulse — Model Comparison ({freq_label})"
        ws2["A3"].font = title_font2
        ws2["A3"].alignment = Alignment(horizontal="left", vertical="center")
        ws2.row_dimensions[3].height = 20.5
        for col_idx in range(1, 7):
            ws2.cell(row=3, column=col_idx).border = thin_border

        # Headers (row 5): Date | Actual | AutoETS | AutoARIMA | Moving Average (Excel) | ETS (Excel)
        comp_headers = ["Date", "Actual"] + MODEL_ORDER
        for col_idx, hdr in enumerate(comp_headers, 1):
            cell = ws2.cell(row=5, column=col_idx, value=hdr)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        # Historical data rows
        r2 = 6
        num_model_cols = len(MODEL_ORDER)
        for entry in historical_data:
            date_str = _format_date(entry["date"], frequency)
            ws2.cell(row=r2, column=1, value=date_str).font = data_font
            ws2.cell(row=r2, column=1).alignment = data_align
            ws2.cell(row=r2, column=1).border = thin_border

            val_cell = ws2.cell(row=r2, column=2, value=entry["value"])
            val_cell.font = data_font
            val_cell.number_format = number_format
            val_cell.alignment = data_align
            val_cell.border = thin_border

            # Model columns blank for historical
            for col in range(3, 3 + num_model_cols):
                c = ws2.cell(row=r2, column=col, value="")
                c.border = thin_border
            r2 += 1

        # Forecast data rows
        for idx, entry in enumerate(forecast_data):
            date_str = _format_date(entry["date"], frequency)
            ws2.cell(row=r2, column=1, value=date_str).font = data_font
            ws2.cell(row=r2, column=1).alignment = data_align
            ws2.cell(row=r2, column=1).border = thin_border

            # Actual column blank
            ws2.cell(row=r2, column=2, value="").border = thin_border

            # Fill each model's forecast value
            for mi, model_name in enumerate(MODEL_ORDER):
                col_idx = 3 + mi
                vals = comparison_forecasts.get(model_name, [])
                val = vals[idx] if idx < len(vals) else None
                cell = ws2.cell(row=r2, column=col_idx, value=val)
                cell.font = data_font
                cell.number_format = number_format
                cell.alignment = data_align
                cell.border = thin_border
            r2 += 1

        # Auto-fit column widths
        for col_idx in range(1, 3 + num_model_cols):
            col_letter = get_column_letter(col_idx)
            max_len = len(comp_headers[col_idx - 1])
            for r in range(6, r2):
                cell = ws2.cell(row=r, column=col_idx)
                val = cell.value
                if val is None or val == "":
                    continue
                if isinstance(val, (int, float)):
                    display_str = f"{val:,.2f}"
                else:
                    display_str = str(val)
                max_len = max(max_len, len(display_str))
            ws2.column_dimensions[col_letter].width = max_len + 4

    # Write to bytes
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
